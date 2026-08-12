#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
三幸町自治会 プログラム配布リスト生成ツール
program_distribution_tool.py

入力:
  - プログラム広告順番案Excel (町外・町内・特別): A列=ID, B列=区部, C列=氏名, G列=寄付金額
  - 積立者リストExcel: A列=ID, B列=区部, C列=氏名

出力列構成:
  A=ID, B=ID漢字, C=ID数字, D=区部, E=名前,
  F=100円券枚数, G=No.つきプログラム, H=No.なしプログラム, I=粗品

配布物ルール:
  - 積立:    F=5固定, G(No.つき)=1, H(No.なし)=0, I(粗品)=0
  - 町外/内:  F=G列÷1000, G=0, H=1, I=1
  - 特別:    F=G列÷1000, G=0, H=0, I=0

その他:
  - 区部ごとに別シート出力（シートタブ名=「X区X部」）
  - 各シートの印刷ヘッダー左側に「X区X部　部長様」を設定
  - 各区部内の並び順: 積立→外→内→特, 各種別内はID番号順
  - データ書き込み後に空白行を削除して表の体裁を保つ
  - B列にIDプレフィックス(漢字)、C列にID番号(数字文字列)を直接書き込む
    例) 「外01」→ B=「外」C=「01」 / 「積立001」→ B=「積立」C=「001」
"""

import os
import re
import sys
import traceback
from datetime import datetime

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------
# バージョン定義 (ここだけ更新すればタイトル・GUI表示に反映される)
# ------------------------------------------------------------------
APP_VERSION = "1.6.0"
APP_TITLE = f"三幸町自治会 プログラム配布リスト生成ツール v{APP_VERSION}"

# ------------------------------------------------------------------
# 列定義
# ------------------------------------------------------------------
# 入力ファイル列
IN_COL_ID     = "A"   # ID
IN_COL_KUBU   = "B"   # 区部
IN_COL_NAME   = "C"   # 氏名
IN_COL_AMOUNT = "G"   # 寄付金額 (町外・町内・特別のみ)

# 出力テンプレート列 (1始まり)
OUT_COL_ID        = 1   # A: ID
OUT_COL_ID_KANJI  = 2   # B: IDプレフィックス(漢字)
OUT_COL_ID_NUM    = 3   # C: ID番号(数字文字列)
OUT_COL_KUBU      = 4   # D: 区部
OUT_COL_NAME      = 5   # E: 名前
OUT_COL_KEN       = 6   # F: 100円券枚数
OUT_COL_NO_TSUKI  = 7   # G: No.つきプログラム (積立=1, 他=0)
OUT_COL_NO_NASHI  = 8   # H: No.なしプログラム (町外/内=1, 他=0)
OUT_COL_SOSHIN    = 9   # I: 粗品 (町外/内=1, 他=0)
OUT_COL_MAX       = 9   # 最終列

DATA_START_ROW = 3   # 1行目=合計行, 2行目=列ヘッダー, 3行目〜=データ

# ID分類パターン: 先頭の漢字部分と末尾の数字部分を分離
# 先頭ゼロは番号としては除去しないが、文字列(C列)はそのまま保持する
ID_PATTERN = re.compile(r"^\s*([^\d]+?)\s*(\d+)\s*$")

# 種別の並び順 (積立=0, 外=1, 内=2, 特=3)
CATEGORY_SORT_ORDER = {
    "積立": 0,
    "外":   1,
    "内":   2,
    "特":   3,
}

# 積立1件あたりの寄付相当金額(集計用)
TSUMITATE_UNIT_AMOUNT = 6000


# ------------------------------------------------------------------
# コアロジック
# ------------------------------------------------------------------

def parse_id(id_raw):
    """
    IDから (プレフィックス文字列, 番号整数, 番号文字列) を返す。
    例: 「外01」→ (「外」, 1, 「01」)
        「積立001」→ (「積立」, 1, 「001」)
        「内04」→ (「内」, 4, 「04」)
    解析不能なら None。
    """
    if id_raw is None:
        return None
    m = ID_PATTERN.match(str(id_raw).strip())
    if not m:
        return None
    prefix = m.group(1)
    num_str = m.group(2)   # 先頭ゼロを保持した文字列 例:「01」「001」
    try:
        number = int(num_str)
    except ValueError:
        return None
    return prefix, number, num_str


def normalize_prefix(prefix):
    """プレフィックスの表記ゆれを正規化する。"""
    mapping = {
        "積立": "積立",
        "外":   "外",
        "内":   "内",
        "特":   "特",
        "得":   "特",   # 表記ゆれ
        "特別": "特",   # 表記ゆれ
    }
    return mapping.get(prefix, prefix)


def kubu_to_label(kubu_str):
    """
    「X-Y」形式を「X区Y部」に変換する。
    例: 「1-1」→「1区1部」, 「2-3」→「2区3部」
    変換できない場合は元の文字列をそのまま返す。
    """
    if kubu_str is None:
        return ""
    s = str(kubu_str).strip()
    m = re.match(r"^(\d+)-(\d+)$", s)
    if m:
        return f"{m.group(1)}区{m.group(2)}部"
    return s


def load_rows(filepath, sheet_name, source_type):
    """
    Excelファイルからデータを読み込んで辞書のリストを返す。
    source_type: "積立" | "寄付者"
    各行辞書キー: id, prefix, prefix_raw, num_str, number,
                  kubu, name, amount, source_type
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        stripped = sheet_name.strip()
        candidates = [s for s in wb.sheetnames if s.strip() == stripped]
        if len(candidates) == 1:
            sheet_name = candidates[0]
        else:
            available = "、".join(wb.sheetnames)
            raise ValueError(
                f"シート '{sheet_name}' が見つかりません。\n"
                f"シート一覧: {available}"
            )
    ws = wb[sheet_name]
    rows = []
    for row_idx in range(2, ws.max_row + 1):   # 1行目はヘッダー想定
        id_val = ws[f"{IN_COL_ID}{row_idx}"].value
        if id_val is None or str(id_val).strip() == "":
            continue
        parsed = parse_id(id_val)
        if parsed is None:
            continue
        prefix_raw, number, num_str = parsed
        prefix = normalize_prefix(prefix_raw)

        kubu_val = ws[f"{IN_COL_KUBU}{row_idx}"].value
        kubu = str(kubu_val).strip() if kubu_val is not None else ""

        name_val = ws[f"{IN_COL_NAME}{row_idx}"].value
        name = str(name_val).strip() if name_val is not None else ""

        amount = 0.0
        if source_type != "積立":
            amount_val = ws[f"{IN_COL_AMOUNT}{row_idx}"].value
            try:
                amount = float(amount_val) if amount_val is not None else 0.0
            except (TypeError, ValueError):
                amount = 0.0

        rows.append({
            "id":         str(id_val).strip(),
            "prefix":     prefix,
            "prefix_raw": prefix_raw,  # B列に書くプレフィックス文字列
            "num_str":    num_str,      # C列に書く番号文字列(先頭ゼロ保持)
            "number":     number,
            "kubu":       kubu,
            "name":       name,
            "amount":     amount,
            "source_type": source_type,
        })
    wb.close()
    return rows


def compute_distribution(row):
    """
    配布物数量を計算する。
    戻り値: (100円券枚数, No.つき, No.なし, 粗品)
      積立:    (5, 1, 0, 0)
      町外/内:  (G÷1000, 0, 1, 1)
      特別:    (G÷1000, 0, 0, 0)
    """
    prefix = row["prefix"]
    if prefix == "積立":
        return 5, 1, 0, 0
    elif prefix in ("外", "内"):
        ken = int(row["amount"] / 1000) if row["amount"] else 0
        return ken, 0, 1, 1
    elif prefix == "特":
        ken = int(row["amount"] / 1000) if row["amount"] else 0
        return ken, 0, 0, 0
    else:
        return 0, 0, 0, 0


def sort_rows(all_rows):
    """
    区部 → 種別(積立→外→内→特) → ID番号 の順にソートする。
    区部は「X-Y」を数値タプル (X, Y) として自然順ソート。
    """
    def kubu_sort_key(kubu_str):
        m = re.match(r"^(\d+)-(\d+)$", str(kubu_str).strip())
        if m:
            return (int(m.group(1)), int(m.group(2)))
        return (9999, 9999)

    def row_sort_key(r):
        kb = kubu_sort_key(r["kubu"])
        cat = CATEGORY_SORT_ORDER.get(r["prefix"], 99)
        return (kb[0], kb[1], cat, r["number"])

    return sorted(all_rows, key=row_sort_key)


def compute_summary(all_rows):
    """
    種別ごとの件数と金額合計を計算する。
    積立は件数×TSUMITATE_UNIT_AMOUNT円を金額として扱う。
    戻り値: dict { prefix: {"count": int, "amount": float} }
    """
    summary = {}
    for r in all_rows:
        p = r["prefix"]
        if p not in summary:
            summary[p] = {"count": 0, "amount": 0.0}
        summary[p]["count"] += 1
        if p == "積立":
            summary[p]["amount"] += TSUMITATE_UNIT_AMOUNT
        else:
            summary[p]["amount"] += r["amount"]
    return summary


def write_output(template_path, output_path, all_rows):
    """
    テンプレートをベースに出力Excelを生成する。

    ・区部ごとに別シートに分けて出力する（シートタブ名=「X区X部」）
    ・各シートの印刷ヘッダー左側に「X区X部　部長様」を設定する
      （Excelの「ページレイアウト→ヘッダー/フッター」左側に表示）
    ・データ行へのヘッダー行挿入は行わない
    ・B列にIDプレフィックス(漢字)、C列にID番号(数字文字列)を直接書き込む
      (数式ではなく値として書き込むことで先頭ゼロの欠落を防ぐ)
    """
    wb = openpyxl.load_workbook(template_path)
    wb.template = False

    # テンプレートのスタイルを元シートから取得しておく
    ws_tmpl = wb.active
    base_font_name = "游ゴシック"
    try:
        fn = ws_tmpl.cell(row=3, column=1).font
        if fn and fn.name:
            base_font_name = fn.name
    except Exception:
        pass

    # データ行フォント
    data_font = Font(name=base_font_name)

    sorted_data = sort_rows(all_rows)

    # 区部ごとにデータをグループ化
    from collections import OrderedDict
    kubu_groups = OrderedDict()
    for r in sorted_data:
        kb = r["kubu"]
        if kb not in kubu_groups:
            kubu_groups[kb] = []
        kubu_groups[kb].append(r)

    # テンプレートのアクティブシートを最初の区部用として使い、
    # 残りは複製して追加する
    first_kubu = True
    all_sheets = []

    for kubu, rows_in_kubu in kubu_groups.items():
        label = kubu_to_label(kubu)

        if first_kubu:
            ws = wb.active
            first_kubu = False
        else:
            # テンプレートシートを複製
            ws = wb.copy_worksheet(wb.worksheets[0])

        ws.title = label   # シートタブ名を「X区X部」にする

        # シート表示をページレイアウト表示に設定
        ws.sheet_view.view = "pageLayout"

        # 印刷ヘッダー: 左側に「X区X部　部長様」、中央に「プログラム・金券・記念品配布一覧」
        ws.oddHeader.left.text = f"{label}　部長様"
        ws.oddHeader.left.size = 14
        ws.oddHeader.left.font = "游ゴシック,Bold"
        ws.oddHeader.center.text = "プログラム・金券・記念品配布一覧"
        ws.oddHeader.center.size = 14
        ws.oddHeader.center.font = "游ゴシック,Bold"

        # フッター中央: ページ番号 (&P)
        ws.oddFooter.center.text = "&P"

        # タイトル行: 2行目を全ページに繰り返し印刷
        ws.print_title_rows = "$2:$2"

        # 枠線を印刷する
        ws.print_options.gridLines = True
        ws.print_options.gridLinesSet = True

        # データ領域をクリア (DATA_START_ROW以降)
        max_r = ws.max_row
        for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=max_r):
            for cell in row:
                cell.value = None
        # 結合セルの解除 (DATA_START_ROW以降)
        for merge_range in list(ws.merged_cells.ranges):
            if merge_range.min_row >= DATA_START_ROW:
                ws.merged_cells.remove(merge_range)

        write_row = DATA_START_ROW

        # データ行を書き込む
        for row_data in rows_in_kubu:
            ken, no_tsuki, no_nashi, soshin = compute_distribution(row_data)

            ws.cell(row=write_row, column=OUT_COL_ID,
                    value=row_data["id"]).font = data_font
            # B列: IDの漢字プレフィックスを直接書き込む (数式不使用)
            ws.cell(row=write_row, column=OUT_COL_ID_KANJI,
                    value=row_data["prefix_raw"]).font = data_font
            # C列: ID番号文字列を直接書き込む (先頭ゼロ保持のため文字列型で設定)
            c_cell = ws.cell(row=write_row, column=OUT_COL_ID_NUM)
            c_cell.value = row_data["num_str"]
            c_cell.number_format = "@"   # 文字列書式
            c_cell.font = data_font
            ws.cell(row=write_row, column=OUT_COL_KUBU,
                    value=row_data["kubu"]).font = data_font
            ws.cell(row=write_row, column=OUT_COL_NAME,
                    value=row_data["name"]).font = data_font
            ws.cell(row=write_row, column=OUT_COL_KEN,
                    value=ken).font = data_font
            ws.cell(row=write_row, column=OUT_COL_NO_TSUKI,
                    value=no_tsuki).font = data_font
            ws.cell(row=write_row, column=OUT_COL_NO_NASHI,
                    value=no_nashi).font = data_font
            ws.cell(row=write_row, column=OUT_COL_SOSHIN,
                    value=soshin).font = data_font

            write_row += 1

        last_data_row = write_row - 1

        # データ書き込み後の空白行を削除して表の体裁を保つ
        # (テンプレートにあらかじめ用意された空白行を末尾から削除)
        rows_to_delete = []
        for r in range(last_data_row + 1, max_r + 1):
            if all(ws.cell(row=r, column=c).value is None
                   for c in range(1, OUT_COL_MAX + 1)):
                rows_to_delete.append(r)
        # 末尾から削除することで行番号ずれを防ぐ
        for r in reversed(rows_to_delete):
            ws.delete_rows(r)

        # 1行目の合計SUBTOTAL式を実データ範囲に更新
        ws["F1"] = f"=SUBTOTAL(9,F{DATA_START_ROW}:F{last_data_row})"
        ws["G1"] = f"=SUBTOTAL(9,G{DATA_START_ROW}:G{last_data_row})"
        ws["H1"] = f"=SUBTOTAL(9,H{DATA_START_ROW}:H{last_data_row})"
        ws["I1"] = f"=SUBTOTAL(9,I{DATA_START_ROW}:I{last_data_row})"

        # 印刷範囲: A1〜I列のデータ最終行まで (last_data_row確定後に設定)
        ws.print_area = f"A1:I{last_data_row}"

        all_sheets.append(ws)

    wb.save(output_path)
    return len(sorted_data)


# ------------------------------------------------------------------
# GUI
# ------------------------------------------------------------------

class FileSelector(ttk.LabelFrame):
    """Excelファイル選択 + シート選択"""

    def __init__(self, master, label_text):
        super().__init__(master, text=label_text, padding=8)
        self.filepath_var = tk.StringVar()
        self.sheet_var = tk.StringVar()

        ttk.Entry(self, textvariable=self.filepath_var, width=55,
                  state="readonly").grid(row=0, column=0, padx=(0, 6), sticky="we")
        ttk.Button(self, text="ファイル選択...",
                   command=self.browse_file).grid(row=0, column=1, sticky="e")

        sheet_row = ttk.Frame(self)
        sheet_row.grid(row=1, column=0, columnspan=2, sticky="w", pady=(6, 0))
        ttk.Label(sheet_row, text="シート:").pack(side="left")
        self.sheet_combo = ttk.Combobox(
            sheet_row, textvariable=self.sheet_var,
            state="readonly", width=30
        )
        self.sheet_combo.pack(side="left", padx=(6, 0))

        self.columnconfigure(0, weight=1)

    def browse_file(self):
        path = filedialog.askopenfilename(
            title="Excelファイルを選択",
            filetypes=[
                ("Excelファイル", "*.xlsx"),
                ("すべてのファイル", "*.*"),
            ],
        )
        if not path:
            return
        self.filepath_var.set(path)
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            messagebox.showerror("エラー", f"ファイルを開けませんでした:\n{e}")
            return
        self.sheet_combo["values"] = sheet_names
        if sheet_names:
            self.sheet_combo.current(0)

    def get_filepath(self):
        return self.filepath_var.get().strip()

    def get_sheet(self):
        return self.sheet_var.get()


class TemplateSelector(ttk.LabelFrame):
    """テンプレートファイル(.xltx)選択"""

    def __init__(self, master, label_text):
        super().__init__(master, text=label_text, padding=8)
        self.filepath_var = tk.StringVar()

        ttk.Entry(self, textvariable=self.filepath_var, width=55,
                  state="readonly").grid(row=0, column=0, padx=(0, 6), sticky="we")
        ttk.Button(self, text="ファイル選択...",
                   command=self.browse_file).grid(row=0, column=1)
        self.columnconfigure(0, weight=1)

    def browse_file(self):
        path = filedialog.askopenfilename(
            title="テンプレートExcelファイルを選択",
            filetypes=[
                ("Excelテンプレート", "*.xltx"),
                ("すべてのファイル", "*.*"),
            ],
        )
        if path:
            self.filepath_var.set(path)

    def get_filepath(self):
        return self.filepath_var.get().strip()


class ProgramDistributionApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)

        # アイコン設定（PyInstaller --onefile 対応）
        try:
            if hasattr(sys, "_MEIPASS"):
                icon_path = os.path.join(sys._MEIPASS, "app_icon.ico")
            else:
                icon_path = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)), "app_icon.ico"
                )
            self.iconbitmap(icon_path)
        except Exception:
            pass

        screen_h = self.winfo_screenheight()
        init_height = min(720, max(480, screen_h - 120))
        self.geometry(f"740x{init_height}")
        self.minsize(620, 420)
        self.resizable(True, True)

        # --- 下部固定バー: 生成ボタン + バージョン表示 ---
        bottom_bar = ttk.Frame(self, padding=(12, 8))
        bottom_bar.pack(side="bottom", fill="x")
        ttk.Button(bottom_bar, text="生成",
                   command=self.on_generate).pack(side="left")
        # APP_VERSION を参照してバージョン表示 (APP_VERSION を変えるだけで反映)
        # ttk.Label を使うことでGC問題を回避し確実に表示する
        ttk.Label(
            bottom_bar,
            text=f"v{APP_VERSION}",
            relief="sunken",
            padding=(4, 2),
        ).pack(side="left", padx=(8, 0))

        # --- スクロール可能なメインエリア ---
        scroll_container = ttk.Frame(self)
        scroll_container.pack(side="top", fill="both", expand=True)

        canvas = tk.Canvas(scroll_container, borderwidth=0, highlightthickness=0)
        vscroll = ttk.Scrollbar(scroll_container, orient="vertical",
                                command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        vscroll.pack(side="right", fill="y")

        main = ttk.Frame(canvas, padding=12)
        canvas_window = canvas.create_window((0, 0), window=main, anchor="nw")

        def _on_frame_configure(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)

        main.bind("<Configure>", _on_frame_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # 説明ラベル
        ttk.Label(
            main,
            text="プログラム広告順番Excelファイル・積立者リストExcel・テンプレートを選択して「生成」を押してください。",
            wraplength=700,
        ).pack(anchor="w", pady=(0, 10))

        # プログラム広告順番案ファイル選択
        self.program_selector = FileSelector(main, "プログラム広告順番Excelファイル")
        self.program_selector.pack(fill="x", pady=4)

        # 積立者リスト選択
        self.tsumitate_selector = FileSelector(main, "積立者リストExcel")
        self.tsumitate_selector.pack(fill="x", pady=4)

        # テンプレート選択
        self.template_selector = TemplateSelector(main, "出力テンプレートファイル (.xltx)")
        self.template_selector.pack(fill="x", pady=(10, 4))

        # 出力ファイル名
        out_frame = ttk.LabelFrame(main, text="出力ファイル名", padding=8)
        out_frame.pack(fill="x", pady=4)
        default_name = f"プログラム配布リスト_{datetime.now().strftime('%Y%m%d')}.xlsx"
        self.output_name_var = tk.StringVar(value=default_name)
        ttk.Entry(out_frame, textvariable=self.output_name_var,
                  width=55).pack(side="left", fill="x", expand=True)

        # 集計結果
        summary_frame = ttk.LabelFrame(main, text="集計結果", padding=8)
        summary_frame.pack(fill="both", expand=True, pady=(10, 0))
        self.summary_text = tk.Text(summary_frame, height=12,
                                    state="disabled", wrap="word")
        self.summary_text.pack(fill="both", expand=True)

    def log(self, text, clear=False):
        self.summary_text.configure(state="normal")
        if clear:
            self.summary_text.delete("1.0", "end")
        self.summary_text.insert("end", text + "\n")
        self.summary_text.configure(state="disabled")

    def on_generate(self):
        try:
            # --- 入力チェック ---
            prog_path  = self.program_selector.get_filepath()
            prog_sheet = self.program_selector.get_sheet()
            tsumi_path  = self.tsumitate_selector.get_filepath()
            tsumi_sheet = self.tsumitate_selector.get_sheet()
            tmpl_path  = self.template_selector.get_filepath()

            if not prog_path or not prog_sheet:
                messagebox.showwarning(
                    "入力不足",
                    "プログラム広告順番ExcelのファイルとシートをどちらもEしてください。")
                return
            if not tsumi_path or not tsumi_sheet:
                messagebox.showwarning(
                    "入力不足", "積立者リストExcelのファイルとシートを選択してください。")
                return
            if not tmpl_path:
                messagebox.showwarning(
                    "入力不足", "テンプレートファイルを選択してください。")
                return

            # --- データ読み込み ---
            try:
                prog_rows = load_rows(prog_path, prog_sheet, "寄付者")
            except Exception as e:
                messagebox.showerror(
                    "エラー",
                    f"プログラム広告順番Excelの読み込みでエラーが発生しました:\n{e}")
                return

            try:
                tsumi_rows = load_rows(tsumi_path, tsumi_sheet, "積立")
            except Exception as e:
                messagebox.showerror(
                    "エラー",
                    f"積立者リストExcelの読み込みでエラーが発生しました:\n{e}")
                return

            all_rows = tsumi_rows + prog_rows

            if not all_rows:
                messagebox.showwarning(
                    "データなし", "有効なデータが読み込めませんでした。")
                return

            # --- 出力先 ---
            base_dir = os.path.dirname(os.path.abspath(prog_path))
            output_dir = os.path.join(base_dir, "output")
            os.makedirs(output_dir, exist_ok=True)
            output_name = self.output_name_var.get().strip() or "プログラム配布リスト.xlsx"
            if not output_name.lower().endswith(".xlsx"):
                output_name += ".xlsx"
            output_path = os.path.join(output_dir, output_name)

            # --- 生成 ---
            total = write_output(tmpl_path, output_path, all_rows)

            # --- 集計表示 ---
            self.log(
                f"=== 生成完了: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===",
                clear=True
            )
            self.log(f"出力先: {output_path}")
            self.log(f"総件数: {total} 件\n")

            # 種別ごとの件数・金額集計
            summary = compute_summary(all_rows)
            grand_total = 0.0
            display_order = [
                ("積立",   "積立",   f"(1件={TSUMITATE_UNIT_AMOUNT:,}円換算)"),
                ("町外(外)", "外",   ""),
                ("町内(内)", "内",   ""),
                ("特別(特)", "特",   ""),
            ]
            self.log("【種別集計】")
            for disp_label, key, note in display_order:
                if key in summary:
                    cnt = summary[key]["count"]
                    amt = summary[key]["amount"]
                    grand_total += amt
                    note_str = f" {note}" if note else ""
                    self.log(
                        f"  {disp_label}: {cnt} 件  {amt:,.0f} 円{note_str}"
                    )
            self.log(f"\n  合計金額: {grand_total:,.0f} 円")

            # 区部ごとの件数
            kubu_counts = {}
            for r in all_rows:
                kb = r["kubu"]
                kubu_counts[kb] = kubu_counts.get(kb, 0) + 1

            self.log("\n【区部ごとの件数】")
            for kb in sorted(
                kubu_counts.keys(),
                key=lambda s: (
                    (int(s.split("-")[0]), int(s.split("-")[1]))
                    if re.match(r"^\d+-\d+$", s) else (9999, 9999)
                )
            ):
                self.log(f"  {kubu_to_label(kb)}: {kubu_counts[kb]} 件")

            messagebox.showinfo("完了",
                                f"生成が完了しました。\n出力先:\n{output_path}")

        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("エラー", f"処理中にエラーが発生しました:\n{e}")


def main():
    app = ProgramDistributionApp()
    app.mainloop()


if __name__ == "__main__":
    main()
